import csv
import io
import re
import uuid
from datetime import datetime, timedelta, timezone

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import case, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.core.deps import CurrentUser, require_admin, require_admin_or_manager, require_org_user
from app.models.call_log import CallLog
from app.models.lead import Lead, LeadCategory, LeadSource, LeadStatus
from app.models.lead_category import LeadCategoryOption
from app.models.lead_assignment import LeadAssignmentHistory
from app.models.user import User, UserRole
from app.schemas.lead import (
    AssignmentHistoryOut,
    AutoAssignResult,
    BulkImportResult,
    BulkReassignRequest,
    BulkReassignResult,
    DuplicateLeadMatch,
    LeadCategoryCreate,
    LeadCategoryOptionOut,
    LeadCreate,
    LeadOut,
    LeadUpdate,
    LastCallOut,
    PaginatedLeads,
    ReassignRequest,
)
from app.schemas.lost_deal import MarkLostRequest
from app.services.distribution import assign_batch, assign_next_telecaller
from app.services.assignment_history import record_assignment

router = APIRouter(prefix="/leads", tags=["leads"])

MAX_IMPORT_FILE_SIZE_BYTES = 10 * 1024 * 1024
MAX_RETURNED_IMPORT_ISSUES = 100
DEFAULT_CATEGORY_BY_KEY = {category.value: category for category in LeadCategory}

LEAD_LOAD_OPTIONS = (
    selectinload(Lead.assignee),
    selectinload(Lead.call_logs),
)


def _to_out(lead: Lead) -> LeadOut:
    out = LeadOut.model_validate(lead)
    out.category = lead.custom_category or lead.category.value
    out.interested_categories = lead.interested_categories or [out.category]
    out.assignee_name = lead.assignee.name if lead.assignee else None
    if lead.call_logs:
        latest = lead.call_logs[0]
        out.last_call = LastCallOut(
            outcome=latest.outcome,
            duration_minutes=float(latest.duration_minutes),
            notes=latest.notes,
            created_at=latest.created_at,
        )
    return out


def _assignment_to_out(event: LeadAssignmentHistory) -> AssignmentHistoryOut:
    return AssignmentHistoryOut(
        id=event.id,
        previous_assignee_id=event.previous_assignee_id,
        previous_assignee_name=event.previous_assignee.name if event.previous_assignee else None,
        new_assignee_id=event.new_assignee_id,
        new_assignee_name=event.new_assignee.name if event.new_assignee else None,
        assigned_by_id=event.assigned_by_id,
        assigned_by_name=event.assigned_by.name if event.assigned_by else None,
        action=event.action,
        source=event.source,
        created_at=event.created_at,
    )


def _category_label(value: str) -> str:
    return value.replace("_", " ").strip().title()


async def _resolve_categories(
    db: AsyncSession,
    organization_id: uuid.UUID,
    values: list[str],
) -> tuple[LeadCategory, str | None, list[str]]:
    normalized_values: list[str] = []
    seen: set[str] = set()
    for value in values:
        normalized = (value or "").strip()
        if not normalized:
            continue
        key = normalized.casefold()
        if key not in seen:
            seen.add(key)
            normalized_values.append(normalized)
    if not normalized_values:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "Select at least one lead category")

    resolved: list[str] = []
    for value in normalized_values:
        default_category = DEFAULT_CATEGORY_BY_KEY.get(value.casefold())
        if default_category is not None:
            resolved.append(default_category.value)
            continue
        result = await db.execute(
            select(LeadCategoryOption).where(
                LeadCategoryOption.organization_id == organization_id,
                func.lower(LeadCategoryOption.name) == value.casefold(),
            )
        )
        custom = result.scalar_one_or_none()
        if custom is None:
            raise HTTPException(
                status.HTTP_422_UNPROCESSABLE_ENTITY,
                f"Category '{value}' is not available in this workspace. Ask an admin to create it first.",
            )
        resolved.append(custom.name)

    first = resolved[0]
    primary = DEFAULT_CATEGORY_BY_KEY.get(first.casefold())
    return primary or LeadCategory.other, None if primary else first, resolved


async def _category_options(db: AsyncSession, organization_id: uuid.UUID) -> list[LeadCategoryOptionOut]:
    result = await db.execute(
        select(LeadCategoryOption.name)
        .where(LeadCategoryOption.organization_id == organization_id)
        .order_by(func.lower(LeadCategoryOption.name))
    )
    options = [
        LeadCategoryOptionOut(value=category.value, label=_category_label(category.value), is_custom=False)
        for category in LeadCategory
    ]
    options.extend(
        LeadCategoryOptionOut(value=name, label=name, is_custom=True)
        for (name,) in result.all()
    )
    return options


def _build_lead_filter_stmt(
    current: CurrentUser,
    source: LeadSource | None,
    status_filter: LeadStatus | None,
    assigned_to: uuid.UUID | None,
    unassigned_only: bool,
    category: str | None,
    state: str | None,
    city: str | None,
    dnd: bool | None,
    q: str | None,
    has_callback: bool | None = None,
):
    stmt = select(Lead).where(Lead.organization_id == current.organization_id)

    if current.role == UserRole.telecaller:
        stmt = stmt.where(Lead.assigned_to == current.id)
    elif assigned_to is not None:
        stmt = stmt.where(Lead.assigned_to == assigned_to)
    if unassigned_only:
        stmt = stmt.where(Lead.assigned_to.is_(None))

    if source is not None:
        stmt = stmt.where(Lead.source == source)
    if status_filter is not None:
        stmt = stmt.where(Lead.status == status_filter)
    if category is not None:
        normalized_category = category.strip().casefold()
        default_category = DEFAULT_CATEGORY_BY_KEY.get(normalized_category)
        stmt = stmt.where(
            Lead.interested_categories.any(default_category.value if default_category is not None else category.strip())
        )
    if state is not None:
        stmt = stmt.where(Lead.state == state)
    if city is not None:
        stmt = stmt.where(Lead.city == city)
    if dnd is not None:
        stmt = stmt.where(Lead.dnd == dnd)
    if has_callback is not None:
        stmt = stmt.where(Lead.next_follow_up_at.isnot(None) if has_callback else Lead.next_follow_up_at.is_(None))
    if q:
        like = f"%{q}%"
        stmt = stmt.where(or_(Lead.name.ilike(like), Lead.phone.ilike(like), Lead.city.ilike(like)))
    return stmt


def _lead_queue_order(current: CurrentUser):
    """Keep a telecaller's next actions at the top of the queue.

    Overdue callbacks are the most time-sensitive work, followed by leads that
    have never been contacted (call pending). Other leads keep the normal
    newest-first ordering used by managers and admins.
    """
    if current.role != UserRole.telecaller:
        return (Lead.created_at.desc(),)

    now = datetime.now(timezone.utc)
    overdue = Lead.next_follow_up_at.isnot(None) & (Lead.next_follow_up_at < now)
    call_pending = Lead.last_contacted_at.is_(None)
    priority = case(
        (overdue, 0),
        (call_pending, 1),
        else_=2,
    )
    overdue_at = case((overdue, Lead.next_follow_up_at), else_=None)
    pending_created_at = case((call_pending, Lead.created_at), else_=None)
    return (
        priority.asc(),
        overdue_at.asc().nullslast(),
        pending_created_at.asc().nullslast(),
        Lead.created_at.desc(),
    )


@router.get("", response_model=PaginatedLeads)
async def list_leads(
    current: CurrentUser = Depends(require_org_user),
    db: AsyncSession = Depends(get_db),
    source: LeadSource | None = None,
    status_filter: LeadStatus | None = Query(default=None, alias="status"),
    assigned_to: uuid.UUID | None = None,
    unassigned_only: bool = False,
    category: str | None = None,
    state: str | None = None,
    city: str | None = None,
    dnd: bool | None = None,
    q: str | None = None,
    has_callback: bool | None = None,
    overdue_only: bool = False,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=200),
):
    stmt = _build_lead_filter_stmt(
        current, source, status_filter, assigned_to, unassigned_only, category, state, city, dnd, q, has_callback
    )
    if overdue_only:
        stmt = stmt.where(Lead.next_follow_up_at.isnot(None), Lead.next_follow_up_at < datetime.now(timezone.utc))

    count_stmt = select(func.count()).select_from(stmt.subquery())
    total = (await db.execute(count_stmt)).scalar_one()

    stmt = (
        stmt.options(*LEAD_LOAD_OPTIONS)
        .order_by(
            *((Lead.next_follow_up_at.asc().nullslast(),) if has_callback else _lead_queue_order(current))
        )
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    result = await db.execute(stmt)
    leads = list(result.scalars().unique().all())

    return PaginatedLeads(items=[_to_out(l) for l in leads], total=total, page=page, page_size=page_size)


@router.get("/export")
async def export_leads(
    current: CurrentUser = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
    source: LeadSource | None = None,
    status_filter: LeadStatus | None = Query(default=None, alias="status"),
    assigned_to: uuid.UUID | None = None,
    category: str | None = None,
    state: str | None = None,
    city: str | None = None,
    dnd: bool | None = None,
    q: str | None = None,
    unassigned_only: bool = False,
):
    stmt = _build_lead_filter_stmt(
        current, source, status_filter, assigned_to, unassigned_only, category, state, city, dnd, q
    )
    stmt = stmt.options(*LEAD_LOAD_OPTIONS).order_by(Lead.created_at.desc())
    result = await db.execute(stmt)
    leads = list(result.scalars().unique().all())

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([
        "Name", "Phone", "City", "State", "Category", "Categories of Interest", "Specialty", "Drug License",
        "Source", "Status", "Assigned To",
        "Credit Limit", "Outstanding", "DND", "Notes", "Created At", "Last Contacted At",
    ])
    for lead in leads:
        writer.writerow([
            lead.name, lead.phone, lead.city or "", lead.state or "", lead.custom_category or lead.category.value,
            ", ".join(lead.interested_categories or [lead.custom_category or lead.category.value]), lead.specialty or "",
            lead.drug_license_number or "",
            lead.source.value, lead.status.value,
            lead.assignee.name if lead.assignee else "", lead.credit_limit or "", lead.outstanding_amount or "",
            "Yes" if lead.dnd else "No", lead.notes or "", lead.created_at.isoformat(),
            lead.last_contacted_at.isoformat() if lead.last_contacted_at else "",
        ])
    buffer.seek(0)

    filename = f"leads-export-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.csv"
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("", response_model=LeadOut, status_code=status.HTTP_201_CREATED)
async def add_lead(
    payload: LeadCreate,
    current: CurrentUser = Depends(require_admin_or_manager),
    db: AsyncSession = Depends(get_db),
):
    category, custom_category, interested_categories = await _resolve_categories(
        db, current.organization_id, payload.interested_categories or [payload.category]
    )
    lead = Lead(
        organization_id=current.organization_id,
        name=payload.name,
        phone=payload.phone,
        city=payload.city,
        state=payload.state,
        source=payload.source,
        notes=payload.notes,
        status=LeadStatus.new,
        category=category,
        custom_category=custom_category,
        interested_categories=interested_categories,
        drug_license_number=payload.drug_license_number,
        specialty=payload.specialty,
        credit_limit=payload.credit_limit,
        outstanding_amount=payload.outstanding_amount,
        dnd=payload.dnd,
    )
    assignee = await assign_next_telecaller(db, current.organization_id)
    if assignee:
        lead.assigned_to = assignee.id
    db.add(lead)
    await db.flush()
    record_assignment(
        db,
        organization_id=current.organization_id,
        lead_id=lead.id,
        previous_assignee_id=None,
        new_assignee_id=lead.assigned_to,
        assigned_by_id=current.id,
        action="created",
        source="manual",
    )
    await db.commit()
    result = await db.execute(select(Lead).options(*LEAD_LOAD_OPTIONS).where(Lead.id == lead.id))
    lead = result.scalar_one()
    return _to_out(lead)


@router.get("/check-duplicate", response_model=list[DuplicateLeadMatch])
async def check_duplicate_phone(
    phone: str = Query(min_length=3),
    current: CurrentUser = Depends(require_org_user),
    db: AsyncSession = Depends(get_db),
):
    stmt = (
        select(Lead)
        .options(selectinload(Lead.assignee))
        .where(Lead.organization_id == current.organization_id, Lead.phone == phone)
    )
    if current.role == UserRole.telecaller:
        stmt = stmt.where(Lead.assigned_to == current.id)
    result = await db.execute(stmt)
    matches = result.scalars().all()
    return [
        DuplicateLeadMatch(
            id=lead.id, name=lead.name, phone=lead.phone, status=lead.status,
            assignee_name=lead.assignee.name if lead.assignee else None,
        )
        for lead in matches
    ]


@router.get("/categories", response_model=list[LeadCategoryOptionOut])
async def list_categories(
    current: CurrentUser = Depends(require_org_user),
    db: AsyncSession = Depends(get_db),
):
    return await _category_options(db, current.organization_id)


@router.post("/categories", response_model=LeadCategoryOptionOut, status_code=status.HTTP_201_CREATED)
async def create_category(
    payload: LeadCategoryCreate,
    current: CurrentUser = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    name = payload.name.strip()
    if not name:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "Category name cannot be blank")
    if name.casefold() in DEFAULT_CATEGORY_BY_KEY:
        raise HTTPException(status.HTTP_409_CONFLICT, f"'{name}' is already a built-in category")

    existing = await db.execute(
        select(LeadCategoryOption).where(
            LeadCategoryOption.organization_id == current.organization_id,
            func.lower(LeadCategoryOption.name) == name.casefold(),
        )
    )
    if existing.scalar_one_or_none() is not None:
        raise HTTPException(status.HTTP_409_CONFLICT, f"Category '{name}' already exists")

    category = LeadCategoryOption(organization_id=current.organization_id, name=name)
    db.add(category)
    await db.commit()
    return LeadCategoryOptionOut(value=category.name, label=category.name, is_custom=True)


@router.get("/cities", response_model=list[str])
async def list_used_cities(
    current: CurrentUser = Depends(require_org_user),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(Lead.city).distinct().where(
        Lead.organization_id == current.organization_id,
        Lead.city.isnot(None),
        func.trim(Lead.city) != "",
    )
    if current.role == UserRole.telecaller:
        stmt = stmt.where(Lead.assigned_to == current.id)
    result = await db.execute(stmt)
    return sorted(row[0] for row in result.all())


@router.get("/{lead_id}/assignment-history", response_model=list[AssignmentHistoryOut])
async def get_assignment_history(
    lead_id: uuid.UUID,
    current: CurrentUser = Depends(require_org_user),
    db: AsyncSession = Depends(get_db),
):
    lead_result = await db.execute(
        select(Lead).where(Lead.id == lead_id, Lead.organization_id == current.organization_id)
    )
    lead = lead_result.scalar_one_or_none()
    if lead is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Lead not found")
    if current.role == UserRole.telecaller and lead.assigned_to != current.id:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Not your lead")

    result = await db.execute(
        select(LeadAssignmentHistory)
        .options(
            selectinload(LeadAssignmentHistory.previous_assignee),
            selectinload(LeadAssignmentHistory.new_assignee),
            selectinload(LeadAssignmentHistory.assigned_by),
        )
        .where(
            LeadAssignmentHistory.lead_id == lead_id,
            LeadAssignmentHistory.organization_id == current.organization_id,
        )
        .order_by(LeadAssignmentHistory.created_at.desc())
    )
    return [_assignment_to_out(event) for event in result.scalars().all()]


@router.get("/{lead_id}", response_model=LeadOut)
async def get_lead(
    lead_id: uuid.UUID,
    current: CurrentUser = Depends(require_org_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(Lead)
        .options(*LEAD_LOAD_OPTIONS)
        .where(Lead.id == lead_id, Lead.organization_id == current.organization_id)
    )
    lead = result.scalar_one_or_none()
    if lead is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Lead not found")
    if current.role == UserRole.telecaller and lead.assigned_to != current.id:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Not your lead")
    return _to_out(lead)


@router.patch("/{lead_id}", response_model=LeadOut)
async def update_lead(
    lead_id: uuid.UUID,
    payload: LeadUpdate,
    current: CurrentUser = Depends(require_org_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Lead).where(Lead.id == lead_id, Lead.organization_id == current.organization_id))
    lead = result.scalar_one_or_none()
    if lead is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Lead not found")

    data = payload.model_dump(exclude_unset=True)
    if current.role == UserRole.telecaller:
        if lead.assigned_to != current.id:
            raise HTTPException(status.HTTP_403_FORBIDDEN, "Not your lead")
        disallowed = set(data.keys()) - {"status", "notes", "category", "interested_categories"}
        if disallowed:
            raise HTTPException(
                status.HTTP_403_FORBIDDEN,
                "Telecallers can only update status, notes, or interested categories",
            )

    previous_assignee_id = lead.assigned_to
    if "assigned_to" in data and data["assigned_to"] is not None:
        assignee_result = await db.execute(
            select(User).where(
                User.id == data["assigned_to"],
                User.organization_id == current.organization_id,
            )
        )
        if assignee_result.scalar_one_or_none() is None:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Assignee not found in this organization")

    if "interested_categories" in data or "category" in data:
        selected = data.get("interested_categories") or [data.get("category") or lead.category.value]
        category, custom_category, interested_categories = await _resolve_categories(
            db, current.organization_id, selected
        )
        data["category"] = category
        data["custom_category"] = custom_category
        data["interested_categories"] = interested_categories

    for field, value in data.items():
        setattr(lead, field, value)

    if "assigned_to" in data and previous_assignee_id != lead.assigned_to:
        record_assignment(
            db,
            organization_id=current.organization_id,
            lead_id=lead.id,
            previous_assignee_id=previous_assignee_id,
            new_assignee_id=lead.assigned_to,
            assigned_by_id=current.id,
            action="reassigned",
            source="manual",
        )

    await db.commit()
    result = await db.execute(select(Lead).options(*LEAD_LOAD_OPTIONS).where(Lead.id == lead.id))
    lead = result.scalar_one()
    return _to_out(lead)


@router.delete("/{lead_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_lead(
    lead_id: uuid.UUID,
    current: CurrentUser = Depends(require_admin_or_manager),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Lead).where(Lead.id == lead_id, Lead.organization_id == current.organization_id))
    lead = result.scalar_one_or_none()
    if lead is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Lead not found")
    await db.delete(lead)
    await db.commit()


@router.post("/{lead_id}/reassign", response_model=LeadOut)
async def reassign_lead(
    lead_id: uuid.UUID,
    payload: ReassignRequest,
    current: CurrentUser = Depends(require_org_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Lead).where(Lead.id == lead_id, Lead.organization_id == current.organization_id))
    lead = result.scalar_one_or_none()
    if lead is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Lead not found")

    if current.role == UserRole.telecaller:
        if lead.assigned_to != current.id:
            raise HTTPException(status.HTTP_403_FORBIDDEN, "Telecallers can only reassign leads assigned to them")
        if payload.assigned_to is None:
            raise HTTPException(status.HTTP_403_FORBIDDEN, "Telecallers must reassign the lead to an active manager")
        user_result = await db.execute(
            select(User).where(
                User.id == payload.assigned_to,
                User.organization_id == current.organization_id,
                User.role == UserRole.manager,
                User.is_active.is_(True),
            )
        )
        if user_result.scalar_one_or_none() is None:
            raise HTTPException(
                status.HTTP_403_FORBIDDEN,
                "Telecallers can only reassign leads to an active manager in their organization",
            )
    elif payload.assigned_to is not None:
        user_result = await db.execute(
            select(User).where(
                User.id == payload.assigned_to,
                User.organization_id == current.organization_id,
                User.is_active.is_(True),
            )
        )
        if user_result.scalar_one_or_none() is None:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Assignee not found in this organization")

    previous_assignee_id = lead.assigned_to
    lead.assigned_to = payload.assigned_to
    if previous_assignee_id != lead.assigned_to:
        record_assignment(
            db,
            organization_id=current.organization_id,
            lead_id=lead.id,
            previous_assignee_id=previous_assignee_id,
            new_assignee_id=lead.assigned_to,
            assigned_by_id=current.id,
            action="reassigned",
            source="manual",
        )
    await db.commit()
    result = await db.execute(select(Lead).options(*LEAD_LOAD_OPTIONS).where(Lead.id == lead.id))
    lead = result.scalar_one()
    return _to_out(lead)


@router.post("/bulk-reassign", response_model=BulkReassignResult)
async def bulk_reassign_leads(
    payload: BulkReassignRequest,
    current: CurrentUser = Depends(require_admin_or_manager),
    db: AsyncSession = Depends(get_db),
):
    """Assign a selected batch of leads to one active telecaller."""
    target_result = await db.execute(
        select(User).where(
            User.id == payload.assigned_to,
            User.organization_id == current.organization_id,
            User.role == UserRole.telecaller,
            User.is_active.is_(True),
        )
    )
    if target_result.scalar_one_or_none() is None:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Choose an active telecaller in your organization",
        )

    result = await db.execute(
        select(Lead).where(
            Lead.organization_id == current.organization_id,
            Lead.id.in_(payload.lead_ids),
        )
    )
    leads = list(result.scalars().all())
    if len(leads) != len(set(payload.lead_ids)):
        raise HTTPException(status.HTTP_404_NOT_FOUND, "One or more selected leads were not found")

    updated_count = 0
    for lead in leads:
        if lead.assigned_to == payload.assigned_to:
            continue
        previous_assignee_id = lead.assigned_to
        lead.assigned_to = payload.assigned_to
        record_assignment(
            db,
            organization_id=current.organization_id,
            lead_id=lead.id,
            previous_assignee_id=previous_assignee_id,
            new_assignee_id=lead.assigned_to,
            assigned_by_id=current.id,
            action="reassigned",
            source="bulk",
        )
        updated_count += 1

    await db.commit()
    return BulkReassignResult(updated_count=updated_count)


@router.post("/auto-assign-unassigned", response_model=AutoAssignResult)
async def auto_assign_unassigned_leads(
    current: CurrentUser = Depends(require_admin_or_manager),
    db: AsyncSession = Depends(get_db),
):
    """Distribute all unassigned leads across active telecallers round-robin."""
    result = await db.execute(
        select(Lead)
        .where(
            Lead.organization_id == current.organization_id,
            Lead.assigned_to.is_(None),
        )
        .order_by(Lead.created_at.asc(), Lead.id.asc())
        .with_for_update()
    )
    leads = list(result.scalars().all())
    if not leads:
        return AutoAssignResult(assigned_count=0, assignments={})

    assignees = await assign_batch(db, current.organization_id, len(leads))
    if not assignees or any(assignee is None for assignee in assignees):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Add an active telecaller before distributing leads")

    assignment_counts: dict[str, int] = {}
    for lead, assignee in zip(leads, assignees):
        assert assignee is not None
        lead.assigned_to = assignee.id
        assignment_counts[assignee.name] = assignment_counts.get(assignee.name, 0) + 1
        record_assignment(
            db,
            organization_id=current.organization_id,
            lead_id=lead.id,
            previous_assignee_id=None,
            new_assignee_id=assignee.id,
            assigned_by_id=current.id,
            action="auto_assigned",
            source="automatic",
        )

    await db.commit()
    return AutoAssignResult(assigned_count=len(leads), assignments=assignment_counts)


@router.post("/{lead_id}/mark-lost", response_model=LeadOut)
async def mark_lead_lost(
    lead_id: uuid.UUID,
    payload: MarkLostRequest,
    current: CurrentUser = Depends(require_org_user),
    db: AsyncSession = Depends(get_db),
):
    """Record a telecaller's lost-deal reason and route the lead to a manager."""
    result = await db.execute(select(Lead).where(Lead.id == lead_id, Lead.organization_id == current.organization_id))
    lead = result.scalar_one_or_none()
    if lead is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Lead not found")
    if current.role != UserRole.telecaller:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Only telecallers can use the lost-deal handoff")
    if lead.assigned_to != current.id:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "You can only mark your own leads as lost")

    manager_result = await db.execute(
        select(User).where(
            User.id == payload.manager_id,
            User.organization_id == current.organization_id,
            User.role == UserRole.manager,
            User.is_active.is_(True),
        )
    )
    if manager_result.scalar_one_or_none() is None:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Choose an active manager in your organization")

    reason = payload.reason.strip()
    if not reason:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "A reason is required when marking a deal as lost")

    db.add(
        CallLog(
            lead_id=lead.id,
            logged_by=current.id,
            duration_minutes=0,
            outcome=LeadStatus.lost,
            notes=reason,
            next_follow_up_at=None,
        )
    )
    lead.status = LeadStatus.lost
    previous_assignee_id = lead.assigned_to
    lead.assigned_to = payload.manager_id
    if previous_assignee_id != lead.assigned_to:
        record_assignment(
            db,
            organization_id=current.organization_id,
            lead_id=lead.id,
            previous_assignee_id=previous_assignee_id,
            new_assignee_id=lead.assigned_to,
            assigned_by_id=current.id,
            action="lost_handoff",
            source="lost_deal",
        )
    lead.last_contacted_at = datetime.now(timezone.utc)
    lead.next_follow_up_at = None
    await db.commit()
    result = await db.execute(select(Lead).options(*LEAD_LOAD_OPTIONS).where(Lead.id == lead.id))
    lead = result.scalar_one()
    return _to_out(lead)


@router.post("/bulk-import", response_model=BulkImportResult)
async def bulk_import_leads(
    source: LeadSource,
    file: UploadFile,
    current: CurrentUser = Depends(require_admin_or_manager),
    db: AsyncSession = Depends(get_db),
):
    raw = await file.read()
    filename = (file.filename or "").lower()

    def raise_import_error(code: str, message: str, issues: list[dict] | None = None, status_code: int = 422):
        raise HTTPException(
            status_code=status_code,
            detail={"code": code, "message": message, "issues": issues or []},
        )

    if not raw:
        raise_import_error("empty_file", "The selected file is empty. Add at least one header row and one data row.")
    if len(raw) > MAX_IMPORT_FILE_SIZE_BYTES:
        raise_import_error(
            "file_too_large",
            "The file is larger than 10 MB. Split the data into smaller files and try again.",
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
        )

    def cell_text(value: object) -> str:
        # Excel stores phone numbers as numbers surprisingly often. Avoid showing
        # a trailing `.0` for integer-valued cells while preserving leading zeros
        # from CSV files.
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip() if value is not None else ""

    def normalize_header(header: str) -> str:
        normalized = re.sub(r"[\s_-]+", " ", header.strip()).casefold()
        aliases = {
            "full name": "name",
            "lead name": "name",
            "customer name": "name",
            "contact name": "name",
            "mobile": "phone",
            "mobile number": "phone",
            "phone number": "phone",
            "contact number": "phone",
            "telephone": "phone",
        }
        return aliases.get(normalized, normalized)

    def validate_headers(headers: list[str]) -> list[str]:
        normalized = [normalize_header(header) for header in headers]
        duplicate_headers = sorted({header for header in normalized if header and normalized.count(header) > 1})
        if duplicate_headers:
            raise_import_error(
                "duplicate_columns",
                f"The file contains duplicate column(s): {', '.join(duplicate_headers)}. Keep each column only once.",
                [
                    {
                        "row": 1,
                        "field": header,
                        "code": "duplicate_column",
                        "message": f"Column '{header}' appears more than once.",
                    }
                    for header in duplicate_headers
                ],
            )

        missing_columns = [column for column in ("name", "phone") if column not in normalized]
        if missing_columns:
            raise_import_error(
                "missing_columns",
                f"The file is missing required column(s): {', '.join(missing_columns)}.",
                [
                    {
                        "row": 1,
                        "field": column,
                        "code": "missing_column",
                        "message": f"Required column '{column}' was not found in the header row.",
                    }
                    for column in missing_columns
                ],
            )
        return normalized

    def add_row_issue(issues: list[dict], issue_count: list[int], **issue: object) -> None:
        issue_count[0] += 1
        if len(issues) < MAX_RETURNED_IMPORT_ISSUES:
            issues.append(issue)

    rows: list[tuple[int, dict[str, object]]] = []
    if filename.endswith(".xlsx"):
        from openpyxl import load_workbook

        try:
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            row_iterator = ws.iter_rows(values_only=True)
            try:
                header_values = next(row_iterator)
            except StopIteration:
                raise_import_error("empty_file", "The spreadsheet has no rows. Add a header row and data, then try again.")
            headers = validate_headers([cell_text(value) for value in header_values])
            for row_number, values in enumerate(row_iterator, start=2):
                values = list(values)
                if not any(cell_text(value) for value in values):
                    continue
                rows.append((row_number, dict(zip(headers, values))))
            wb.close()
        except HTTPException:
            raise
        except Exception:
            raise_import_error(
                "invalid_xlsx",
                "The Excel file could not be read. Make sure it is a valid .xlsx file and try again.",
            )
    elif filename.endswith(".csv") or file.content_type in {"text/csv", "application/csv"}:
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            raise_import_error(
                "invalid_encoding",
                "The CSV file is not UTF-8 encoded. Save it as CSV UTF-8 and try again.",
            )

        try:
            csv_rows = list(csv.reader(io.StringIO(text), strict=True))
        except csv.Error:
            raise_import_error(
                "malformed_csv",
                "The CSV file has malformed quoting or columns. Open it in a spreadsheet app, repair it, and export it as CSV UTF-8.",
            )
        if not csv_rows:
            raise_import_error("empty_file", "The CSV file has no rows. Add a header row and data, then try again.")

        headers = validate_headers(csv_rows[0])
        for row_number, values in enumerate(csv_rows[1:], start=2):
            if not any(cell_text(value) for value in values):
                continue
            if len(values) > len(headers):
                raise_import_error(
                    "extra_columns",
                    f"Row {row_number} has more values than the header row. Check for an extra comma or add the missing header.",
                    [
                        {
                            "row": row_number,
                            "code": "extra_columns",
                            "message": "This row has more values than the header row.",
                        }
                    ],
                )
            padded_values = values + [None] * (len(headers) - len(values))
            rows.append((row_number, dict(zip(headers, padded_values))))
    else:
        raise_import_error(
            "unsupported_file_type",
            "Unsupported file type. Upload a CSV or XLSX file.",
            status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
        )

    if not rows:
        raise_import_error("no_data_rows", "The file has a header but no data rows. Add at least one lead and try again.")

    custom_category_result = await db.execute(
        select(LeadCategoryOption.name).where(LeadCategoryOption.organization_id == current.organization_id)
    )
    custom_categories = {name.casefold(): name for (name,) in custom_category_result.all()}
    category_values = {c.value for c in LeadCategory}
    issues: list[dict] = []
    issue_count = [0]

    valid_rows = []
    skipped = 0
    seen_phones: set[str] = set()
    for row_number, raw_row in rows:
        name = cell_text(raw_row.get("name"))
        phone = cell_text(raw_row.get("phone"))
        row_invalid = False
        if not name:
            row_invalid = True
            add_row_issue(
                issues,
                issue_count,
                row=row_number,
                field="name",
                code="missing_name",
                message="Name is required for every lead.",
                severity="error",
            )
        elif len(name) > 255:
            row_invalid = True
            add_row_issue(
                issues,
                issue_count,
                row=row_number,
                field="name",
                code="name_too_long",
                message="Name must be 255 characters or fewer.",
                severity="error",
            )
        if not phone:
            row_invalid = True
            add_row_issue(
                issues,
                issue_count,
                row=row_number,
                field="phone",
                code="missing_phone",
                message="Phone is required for every lead.",
                severity="error",
            )
        elif len(phone) < 6 or len(phone) > 20:
            row_invalid = True
            add_row_issue(
                issues,
                issue_count,
                row=row_number,
                field="phone",
                code="invalid_phone",
                message="Phone must be between 6 and 20 characters.",
                severity="error",
                value=phone,
            )
        if row_invalid:
            skipped += 1
            continue
        if phone in seen_phones:
            skipped += 1
            add_row_issue(
                issues,
                issue_count,
                row=row_number,
                field="phone",
                code="duplicate_in_file",
                message="This phone number appears more than once in the uploaded file; the later row was skipped.",
                severity="error",
                value=phone,
            )
            continue
        seen_phones.add(phone)
        city = cell_text(raw_row.get("city")) or None
        if city and len(city) > 255:
            skipped += 1
            add_row_issue(
                issues,
                issue_count,
                row=row_number,
                field="city",
                code="city_too_long",
                message="City must be 255 characters or fewer.",
                severity="error",
            )
            continue
        category_raw = cell_text(raw_row.get("category")).casefold()
        custom_category = None
        if category_raw in category_values:
            category = LeadCategory(category_raw)
        elif category_raw in custom_categories:
            category = LeadCategory.other
            custom_category = custom_categories[category_raw]
        else:
            category = LeadCategory.other
        if category_raw and category_raw not in category_values and category_raw not in custom_categories:
            add_row_issue(
                issues,
                issue_count,
                row=row_number,
                field="category",
                code="unknown_category",
                message=f"Category '{category_raw}' is not recognized; the lead was imported as 'other'.",
                severity="warning",
                value=category_raw,
            )
        valid_rows.append(
            {
                "row": row_number,
                "name": name,
                "phone": phone,
                "city": city,
                "category": category,
                "custom_category": custom_category,
                "interested_categories": [custom_category or category.value],
            }
        )

    if not valid_rows:
        return BulkImportResult(
            imported=0,
            skipped=skipped,
            assignments={},
            issues=issues,
            issue_count=issue_count[0],
            issues_truncated=issue_count[0] > len(issues),
        )

    existing_result = await db.execute(
        select(Lead.phone).where(
            Lead.organization_id == current.organization_id, Lead.phone.in_([r["phone"] for r in valid_rows])
        )
    )
    existing_phones = {p for (p,) in existing_result.all()}
    duplicates_skipped = 0
    fresh_rows = []
    for row in valid_rows:
        if row["phone"] in existing_phones:
            duplicates_skipped += 1
            add_row_issue(
                issues,
                issue_count,
                row=row["row"],
                field="phone",
                code="duplicate_existing",
                message="A lead with this phone number already exists in your workspace; this row was skipped.",
                severity="error",
                value=row["phone"],
            )
            continue
        fresh_rows.append(row)
    valid_rows = fresh_rows

    if not valid_rows:
        return BulkImportResult(
            imported=0,
            skipped=skipped,
            duplicates_skipped=duplicates_skipped,
            assignments={},
            issues=issues,
            issue_count=issue_count[0],
            issues_truncated=issue_count[0] > len(issues),
        )

    assignees = await assign_batch(db, current.organization_id, len(valid_rows))
    assignments_count: dict[str, int] = {}
    new_leads = []
    for row, assignee in zip(valid_rows, assignees):
        lead = Lead(
            organization_id=current.organization_id,
            name=row["name"],
            phone=row["phone"],
            city=row["city"],
            source=source,
            status=LeadStatus.new,
            category=row["category"],
            custom_category=row["custom_category"],
            interested_categories=row["interested_categories"],
            assigned_to=assignee.id if assignee else None,
        )
        new_leads.append(lead)
        if assignee:
            assignments_count[assignee.name] = assignments_count.get(assignee.name, 0) + 1

    db.add_all(new_leads)
    await db.flush()
    for lead in new_leads:
        record_assignment(
            db,
            organization_id=current.organization_id,
            lead_id=lead.id,
            previous_assignee_id=None,
            new_assignee_id=lead.assigned_to,
            assigned_by_id=current.id,
            action="created",
            source="bulk_import",
        )
    await db.commit()

    return BulkImportResult(
        imported=len(new_leads),
        skipped=skipped,
        duplicates_skipped=duplicates_skipped,
        assignments=assignments_count,
        issues=issues,
        issue_count=issue_count[0],
        issues_truncated=issue_count[0] > len(issues),
    )


@router.delete("", status_code=status.HTTP_204_NO_CONTENT)
async def clear_all_leads(
    current: CurrentUser = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Lead).where(Lead.organization_id == current.organization_id))
    leads = result.scalars().all()
    for lead in leads:
        await db.delete(lead)
    await db.commit()
